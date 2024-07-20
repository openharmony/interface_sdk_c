#!/usr/bin/env python
# coding=utf-8
##############################################
# Copyright (c) 2021-2022 Huawei Device Co., Ltd.
# Licensed under the Apache License, Version 2.0 (the "License");
# you may not use this file except in compliance with the License.
# You may obtain a copy of the License at
#
#     http://www.apache.org/licenses/LICENSE-2.0
#
# Unless required by applicable law or agreed to in writing, software
# distributed under the License is distributed on an "AS IS" BASIS,
# WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
# See the License for the specific language governing permissions and
# limitations under the License.
##############################################

import json
import os
import re
import openpyxl
from typedef.parser.parser import DifferApiRegular, DifferApiInfor


def compare_json_file(generate_json, original_json):  # 获取对比结果
    with open(generate_json, 'r', encoding='utf-8') as g_js:
        generate_data_dict = json.load(g_js)
    with open(original_json, 'r') as o_js:
        original_data_dict = json.load(o_js)
    compare_result = []
    generate_data_only = []  # 装解析生成的json独有的数据
    result_api = filter_compare(generate_data_dict)
    for it in result_api:
        result_name = it["name"]
        key = 0
        for item in original_data_dict:
            if item["name"] == result_name:
                key = 1
                compare_result.append(it)
                break
        if key == 0:
            generate_data_only.append(it)
    original_data_only = get_difference_data(compare_result, original_data_dict)  # 获取提供的json独有的数据
    g_js.close()
    o_js.close()
    return compare_result, generate_data_only, original_data_only


def get_difference_data(compare_result, original_data):
    original_data_only = []
    for item in original_data:
        original_name = item["name"]
        key = 0
        for it in compare_result:
            compare_name = it["name"]
            if original_name == compare_name:
                key = 1
                break
        if key == 0:
            original_data_only.append(item)
    return original_data_only


def filter_compare(analytic_data):  # 获取函数和变量
    result_api = []
    for it in analytic_data:
        get_result_api(it, result_api)
    return result_api


def difference_api(api_data: dict):
    api_name = api_data['name']
    differ_infor = DifferApiInfor.THIRD_PARTY_API.value
    closed_pattern = DifferApiRegular.CLOSED_SOURCE_API_REGULAR.value
    open_pattern = DifferApiRegular.OPEN_SOURCE_API_REGULAR.value
    if re.search(closed_pattern, api_name, flags=re.IGNORECASE):
        differ_infor = DifferApiInfor.CLOSED_SOURCE_API.value
    elif re.search(open_pattern, api_name, flags=re.IGNORECASE):
        differ_infor = DifferApiInfor.OPEN_SOURCE_API.value

    return differ_infor


def get_result_api(file_data, result_api):
    if 'children' in file_data:
        for item1 in file_data["children"]:  # 抛开根节点
            if (item1["kind"] == 'FUNCTION_DECL' or item1["kind"] == 'VAR_DECL') and item1["is_extern"]:
                differ_infor = difference_api(item1)
                item1['differ_infor'] = differ_infor
                item = filter_func(item1)
                result_api.append(item)


def get_parm(item, parm):
    if item["parm"]:
        for i in range(len(item["parm"])):
            if item["parm"][i]["kind"] != 'PARM_DECL':
                continue
            else:
                str_parm = '{} {}'.format(item["parm"][i]["type"], item["parm"][i]["name"])
                parm.append(str_parm)
        item["parm"] = parm


def filter_func(item):
    del item["is_extern"]  # 剔除is_extern键值对，过滤后都是extern
    del item["comment"]
    if "type_ref" in list(item.keys()):
        del item["type_ref"]
    if "children" in list(item.keys()):
        del item["children"]

    item["location_path"] = item["location"]["location_path"]
    item["location"] = item["location"]["location_line"]
    if item["kind"] == 'FUNCTION_DECL':
        item["kind"] = '函数类型'
        parm = []  # 装函数参数
        if "parm" in item:
            get_parm(item, parm)
    else:
        item["kind"] = '变量类型'
        del item["is_const"]
    return item


def collated_api_data(api_data: list):
    collated_data_total = []
    for api in api_data:
        api_content = ''
        if 'node_content' in api and 'content' in api['node_content']:
            api_content = api['node_content']['content']
            api_content = api_content.replace('   ', '')
        collated_data = [
            api.get('module_name'),
            api.get('class_name'),
            api.get('name'),
            api_content,
            api.get('kind'),
            api.get('since'),
            api.get('deprecate_since'),
            api.get('syscap'),
            api.get('error_num'),
            api.get('is_system_api'),
            api.get('model_constraint'),
            api.get('permission'),
            api.get('cross_platform'),
            api.get('form'),
            api.get('atomic_service'),
            api.get('decorator'),
            api.get('kit_name'),
            api.get('location_path'),
            api.get('sub_system'),
            api.get('unique_id'),
            api.get('differ_infor')
        ]
        collated_data_total.append(collated_data)
    return collated_data_total


def generate_excel(array, name, generate_json_unique, original_json_unique):
    first_line_infor = ['模块名', '类名', '方法名', '函数', '类型', '起始版本',
                        '废弃版本', 'syscap', '错误码', '是否为系统API', '模型限制',
                        '权限', '是否支持跨平台', '是否支持卡片应用', '是否支持高阶API',
                        '装饰器', 'kit', '文件路径', '子系统', '接口全路径', '开源/闭源/三方库API']
    workbook = openpyxl.Workbook()
    work_sheet1 = workbook.active
    work_sheet1.title = '对比结果'
    work_sheet1.append(first_line_infor)
    array_update = collated_api_data(array)
    write_information_to_worksheet(work_sheet1, array_update)

    work_sheet2 = workbook.create_sheet('生成json独有')
    work_sheet2.append(first_line_infor)
    generate_json_unique_update = collated_api_data(generate_json_unique)
    write_information_to_worksheet(work_sheet2, generate_json_unique_update)

    work_sheet3 = workbook.create_sheet('原有json独有')
    write_original_infor_to_worksheet(work_sheet3, original_json_unique)
    workbook.save(name)


def write_information_to_worksheet(work_sheet, information_data):
    for data in information_data:
        write_data = data[0], data[1], data[2], data[3], data[4], \
                     data[5], data[6], data[7], data[8], data[9], \
                     data[10], data[11], data[12], data[13], data[14], \
                     data[15], data[16], data[17], data[18], data[19], data[20]
        work_sheet.append(write_data)


def write_original_infor_to_worksheet(work_sheet, original_data):
    first_line_infor = ['first_introduced', '名称']
    work_sheet.append(first_line_infor)
    collated_original_total = []
    for element in original_data:
        original_list = []
        if 'first_introduced' in element:
            original_list.append(element['first_introduced'])
        else:
            original_list.append('')
        if 'name' in element:
            original_list.append(element['name'])
        else:
            original_list.append('')
        collated_original_total.append(original_list)

    for collated_element in collated_original_total:
        write_data = collated_element[0], collated_element[1]
        work_sheet.append(write_data)


def del_repetition_value(generate_data_only_list, compare_list):
    data = []
    for item in generate_data_only_list:
        if item not in compare_list:
            data.append(item)
    return data


def get_json_file(generate_json_file, original_json_file):  # 获取生成的json文件
    generate_json_file_path = r'{}'.format(generate_json_file)  # 获取要对比的json文件
    head_name = os.path.splitext(generate_json_file_path)  # 去掉文件名后缀
    head_name = head_name[0] + '.xlsx'  # 加后缀
    compare_result_list = []
    generate_data_only = []
    original_data_only = []
    for item in original_json_file:  # 对比每一个json(目录下的)
        # 对比两个json文件
        result_list_part, generate_data, original_data = compare_json_file(generate_json_file_path, item)
        compare_result_list.extend(result_list_part)
        generate_data_only.extend(generate_data)
        original_data_only.extend(original_data)
    generate_data_only_new = del_repetition_value(generate_data_only, compare_result_list)
    return compare_result_list, head_name, generate_data_only_new, original_data_only  # 返回对比数据，和所需表格名


def get_parser_json_data(generate_json_file_path, parser_data):
    generate_json_file_path = r'{}'.format(generate_json_file_path)  # 获取要对比的json文件
    head_name = os.path.splitext(generate_json_file_path)  # 去掉文件名后缀
    head_name = head_name[0] + '.xlsx'  # 加后缀
    compare_result_list = []
    generate_data_only = filter_compare(parser_data)
    original_data_only = []
    return compare_result_list, head_name, generate_data_only, original_data_only  # 返回对比数据，和所需表格名


def get_api_data(parser_data, excel_file_name):
    generate_json_unique = []
    original_json_unique = []
    api_data_list = filter_compare(parser_data)
    generate_excel(api_data_list, excel_file_name, generate_json_unique, original_json_unique)
