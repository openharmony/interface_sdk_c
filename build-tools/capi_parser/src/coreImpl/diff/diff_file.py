#!/usr/bin/env python
# -*- coding: utf-8 -*-
# Copyright (c) 2023 Huawei Device Co., Ltd.
# Licensed under the Apache License, Version 2.0 (the "License");
# you may not use this file except in compliance with the License.
# You may obtain a copy of the License at
#
#     http://www.apache.org/licenses/LICENSE-2.0
#
# Unless required by applicable law or agreed to in writing, software
# distributed under the License is distributed on an "AS IS" BASIS,
# WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
# See the License for the specific language governing permissions and
# limitations under the License.

import filecmp
import json
import os
import stat
import re
from collections import OrderedDict
import openpyxl as op
from coreImpl.parser.parser import diff_parser_include_ast
from coreImpl.diff.diff_processor_node import judgment_entrance, change_data_total
from typedef.diff.diff import OutputJson, ApiChangeData, IgnoreFileDirectory
from bin.write_md import write_md_entrance

global_old_dir = ''
global_new_dir = ''
diff_info_list = []
syntax_file_list = []


def get_modification_type_dict():
    modification_type_dict = {
        'API新增': 0,
        'API删除': 0,
        'API废弃': 0,
        'API修改': 0,
        'API修改（原型修改）': 0,
        'API修改（约束变化）': 0
    }
    return modification_type_dict


def get_compatible_dict():
    compatible_dict = {
        '兼容性': 0,
        '非兼容性': 0
    }
    return compatible_dict


def change_to_json(data):
    data_of_json = json.dumps(data, ensure_ascii=False, indent=4)
    return data_of_json


def get_api_change_obj(api_data):
    modification_type_dict = get_modification_type_dict()
    compatible_dict = get_compatible_dict()
    change_data_obj = ApiChangeData()
    key = 0
    api_unique_id = ''
    for element in api_data:
        api_unique_id = element.current_api_unique_id
        if 0 == key:
            change_data_obj.set_api_name(element.api_node_name)
            change_data_obj.set_kit_name(element.kit_name)
            change_data_obj.set_sub_system(element.sub_system)
            change_data_obj.set_is_api_change(element.is_api_change)
            change_data_obj.set_current_api_type(element.current_api_type)
            change_data_obj.set_diff_type(element.operation_diff_type)
            change_data_obj.set_change_type(element.api_modification_type)
            change_data_obj.set_old_all_text(element.old_api_full_text)
            change_data_obj.set_new_all_text(element.new_api_full_text)
            change_data_obj.set_compatible_total(element.is_compatible)
            change_data_obj.set_is_system_api(element.is_system_api)
            change_data_obj.set_open_close_api(element.open_close_api)
            change_data_obj.set_is_third_party_api(element.is_third_party_api)
            key = 1
        else:
            old_all_text = '{}#&#{}'.format(change_data_obj.old_all_text, element.old_api_full_text)
            new_all_text = '{}#&#{}'.format(change_data_obj.new_all_text, element.new_api_full_text)
            diff_type_all = '{}#&#{}'.format(change_data_obj.get_diff_type(), element.operation_diff_type)
            change_type_all = '{}#&#{}'.format(change_data_obj.get_change_type(), element.api_modification_type)
            compatible_data_all = '{}#&#{}'.format(change_data_obj.get_compatible_total(), element.is_compatible)
            change_data_obj.set_old_all_text(old_all_text)
            change_data_obj.set_new_all_text(new_all_text)
            change_data_obj.set_diff_type(diff_type_all)
            change_data_obj.set_change_type(change_type_all)
            change_data_obj.set_compatible_total(compatible_data_all)
        if element.is_compatible and (0 == compatible_dict.get('兼容性')):
            compatible_dict['兼容性'] = 1
        elif not element.is_compatible and (0 == compatible_dict.get('非兼容性')):
            compatible_dict['非兼容性'] = 1
        if element.api_modification_type in modification_type_dict:
            modification_type_dict[element.api_modification_type] = 1
    if 1 == modification_type_dict.get('API修改（原型修改）') and 1 == modification_type_dict.get('API修改（约束变化）'):
        modification_type_dict['API修改'] = 1
    compatible_str = change_to_json(compatible_dict)
    modification_type_str = change_to_json(modification_type_dict)
    change_data_obj.set_compatible(compatible_str)
    change_data_obj.set_change_num(modification_type_str)
    change_data_obj.set_unique_id(api_unique_id)

    return change_data_obj


def collect_api_change(change_data: list):
    api_change_data = []
    for list_element in change_data:
        change_obj = get_api_change_obj(list_element)
        api_change_data.append(change_obj)

    return api_change_data


def collect_node_api_change(api_change_info_list):
    change_data = []
    for api_change_info in api_change_info_list:
        info_data = [
            api_change_info.api_name,
            api_change_info.kit_name,
            api_change_info.sub_system,
            api_change_info.is_api_change,
            api_change_info.current_api_type,
            api_change_info.diff_type,
            api_change_info.change_type,
            api_change_info.compatible,
            api_change_info.change_num,
            api_change_info.old_all_text,
            api_change_info.new_all_text,
            api_change_info.compatible_total,
            api_change_info.unique_id,
            api_change_info.is_system_api,
            api_change_info.open_close_api,
            api_change_info.is_third_party_api
        ]
        change_data.append(info_data)

    return change_data


def syntax_file_excel(output_path):
    data = []
    if syntax_file_list:
        for syntax_dict in syntax_file_list:
            info_data = [
                syntax_dict.get('current_file'),
                syntax_dict.get('error_message')
            ]
            data.append(info_data)

        wb = op.Workbook()
        ws = wb['Sheet']
        ws.title = '语法错误文件信息'
        ws.append(['当前文件', '错误信息'])
        for element in data:
            d = element[0], element[1]
            ws.append(d)
        output_path_xlsx = os.path.abspath(os.path.join(output_path, r'syntax_file_error.xlsx'))
        wb.save(output_path_xlsx)


def start_diff_file(old_dir, new_dir, output_path):
    result_info_list = global_assignment(old_dir, new_dir)
    total = change_data_total
    collect_api_change_data = collect_api_change(total)
    generate_excel(result_info_list, collect_api_change_data, output_path)
    syntax_file_excel(output_path)
    write_md_entrance(result_info_list, output_path)
    result_json = result_to_json(result_info_list)
    diff_result_path = r'./diff_result.txt'
    output_path_txt = os.path.abspath(os.path.join(output_path, diff_result_path))
    write_in_txt(result_json, output_path_txt)


def check_diff_entrance(old_dir, new_dir):
    result_info_list = global_assignment(old_dir, new_dir)

    return result_info_list


def disposal_result_data(result_info_list):
    data = []
    for diff_info in result_info_list:
        info_data = [
            diff_info.operation_diff_type,
            diff_info.old_api_full_text,
            diff_info.new_api_full_text,
            diff_info.api_file_path,
            diff_info.sub_system,
            diff_info.kit_name,
            diff_info.is_system_api
        ]
        data.append(info_data)

    return data


def generate_excel(result_info_list, api_change_data, output_path):
    data = disposal_result_data(result_info_list)
    wb = op.Workbook()
    ws = wb['Sheet']
    ws.title = 'api差异'
    ws.append(['操作标记', '差异项-旧版本', '差异项-新版本', '.h文件', '归属子系统', 'kit', '是否为系统API'])
    for title in data:
        d = title[0], title[1], title[2], title[3], title[4], title[5], title[6]
        ws.append(d)

    change_data_list = collect_node_api_change(api_change_data)
    ws_of_change = wb.create_sheet('api变更次数统计')
    ws_of_change.append(['api名称', 'kit名称', '归属子系统', '是否是api', 'api类型', '操作标记', '变更类型',
                         '兼容性', '变更次数', '差异项-旧版本', '差异项-新版本', '兼容性列表', '接口全路径',
                         '是否为系统API', '开源/闭源API', '是否是三方库api'])
    for element in change_data_list:
        change_data = element[0], element[1], element[2], element[3], element[4], element[5],\
                      element[6], element[7], element[8], element[9], element[10], element[11],\
                      element[12], element[13], element[14], element[15]
        ws_of_change.append(change_data)
    output_path_xlsx = os.path.abspath(os.path.join(output_path, 'diff.xlsx'))
    wb.save(output_path_xlsx)


def global_assignment(old_dir, new_dir):
    global diff_info_list
    diff_info_list = []
    global global_old_dir
    global_old_dir = old_dir
    global global_new_dir
    global_new_dir = new_dir
    do_diff(old_dir, new_dir)
    return diff_info_list


def result_to_json(result_info_list):
    result_json = []
    for diff_info in result_info_list:
        result_json.append(OutputJson(diff_info))
    return json.dumps(result_json, default=lambda obj: obj.__dict__, indent=4, ensure_ascii=False)


def write_in_txt(check_result, output_path):
    modes = stat.S_IRWXO | stat.S_IRWXG | stat.S_IRWXU
    fd = os.open(output_path, os.O_WRONLY | os.O_CREAT | os.O_TRUNC, mode=modes)
    os.write(fd, check_result.encode())
    os.close(fd)


def do_diff(old_dir, new_dir):
    old_file_list = os.listdir(old_dir)
    new_file_list = os.listdir(new_dir)
    diff_list(old_file_list, new_file_list, old_dir, new_dir)


def get_file_ext(file_name):
    return os.path.splitext(file_name)[1]


def filter_ignore_file(file_path):
    norm_file_path = os.path.normpath(file_path)
    if os.name == 'nt':  # Windows
        pattern = re.compile(IgnoreFileDirectory.IGNORE_FILE_DIR_wd.value)
    else:  # Linux / macOS
        pattern = re.compile(IgnoreFileDirectory.IGNORE_FILE_DIR_lx.value)
    # 检查匹配
    if pattern.search(norm_file_path):
        return False
    return True


def diff_list(old_file_list, new_file_list, old_dir, new_dir):
    all_list = set(old_file_list + new_file_list)
    if len(all_list) == 0:
        return
    for target_file in all_list:
        if (get_file_ext(target_file) != '.h'
                and get_file_ext(target_file) != ''):
            continue
        if (target_file in old_file_list
                and target_file not in new_file_list):
            diff_file_path = os.path.join(old_dir, target_file)
            del_old_file(diff_file_path)
        if (target_file in new_file_list
                and target_file not in old_file_list):
            diff_file_path = os.path.join(new_dir, target_file)
            add_new_file(diff_file_path)
        get_same_file_diff(target_file, old_file_list, new_file_list, old_dir, new_dir)


def add_new_file(diff_file_path):
    if os.path.isdir(diff_file_path):
        add_file(diff_file_path)
    elif filter_ignore_file(diff_file_path):
        result_map = parse_file_result(diff_parser_include_ast(global_new_dir, [diff_file_path], flag=1))
        for new_info in result_map.values():
            diff_info_list.extend(judgment_entrance(None, new_info))


def del_old_file(diff_file_path):
    if os.path.isdir(diff_file_path):
        del_file(diff_file_path)
    elif filter_ignore_file(diff_file_path):
        result_map = parse_file_result(diff_parser_include_ast(global_old_dir, [diff_file_path], flag=0))
        for old_info in result_map.values():
            diff_info_list.extend(judgment_entrance(old_info, None))


def get_same_file_diff(target_file, old_file_list, new_file_list, old_dir, new_dir):
    if (target_file in old_file_list
            and target_file in new_file_list):
        if (os.path.isdir(os.path.join(old_dir, target_file))
                and os.path.isdir(os.path.join(new_dir, target_file))):
            old_child_dir = os.path.join(old_dir, target_file)
            new_child_dir = os.path.join(new_dir, target_file)
            do_diff(old_child_dir, new_child_dir)
        if (os.path.isfile(os.path.join(old_dir, target_file))
                and os.path.isfile(os.path.join(new_dir, target_file))):
            old_target_file = os.path.join(old_dir, target_file)
            new_target_file = os.path.join(new_dir, target_file)
            if not filecmp.cmp(old_target_file, new_target_file):
                get_file_result_diff(old_target_file, new_target_file)


def get_file_result_diff(old_target_file, new_target_file):
    if filter_ignore_file(old_target_file):
        old_file_result_map = parse_file_result(diff_parser_include_ast(global_old_dir, [old_target_file], flag=0))
        new_file_result_map = parse_file_result(diff_parser_include_ast(global_new_dir, [new_target_file], flag=1))
        if old_file_result_map and new_file_result_map:
            merged_dict = OrderedDict(list(old_file_result_map.items()) + list(new_file_result_map.items()))
            all_key_list = merged_dict.keys()
            for key in all_key_list:
                diff_info_list.extend(judgment_entrance(old_file_result_map.get(key), new_file_result_map.get(key)))


def del_file(dir_path):
    file_list = os.listdir(dir_path)
    for i in file_list:
        if get_file_ext(i) != '.h' and get_file_ext(i) != '':
            continue
        file_path = os.path.join(dir_path, i)
        if os.path.isdir(file_path):
            del_file(file_path)
        if get_file_ext(i) == '.h' and filter_ignore_file(file_path):
            result_map = parse_file_result(diff_parser_include_ast(global_old_dir, [file_path], flag=0))
            for old_info in result_map.values():
                diff_info_list.extend(judgment_entrance(old_info, None))


def add_file(dir_path):
    file_list = os.listdir(dir_path)
    for i in file_list:
        if get_file_ext(i) != '.h' and get_file_ext(i) != '':
            continue
        file_path = os.path.join(dir_path, i)
        if os.path.isdir(file_path):
            add_file(file_path)
        if get_file_ext(i) == '.h' and filter_ignore_file(file_path):
            result_map = parse_file_result(diff_parser_include_ast(global_new_dir, [file_path], flag=1))
            for new_info in result_map.values():
                diff_info_list.extend(judgment_entrance(None, new_info))


def parse_file_result(result, data_type=0):
    """
    Args:
        result: ***
        data_type(int): 数据处理类型。1-文件新增或删除；0-其他
    """
    result_map = {}
    for root_node in result:
        if root_node['syntax_error'] != 'NA':
            error_file_path = os.path.abspath(os.path.join(root_node['gn_path'], root_node['name']))
            error_message_dict = {
                'current_file': error_file_path,
                'error_message': root_node['syntax_error']
            }
            syntax_file_list.append(error_message_dict)
        result_map.setdefault(f'{root_node["name"]}-{root_node["kind"]}', root_node)
        if data_type != 1:
            parse_file_result_by_child(result_map, root_node)
    return result_map


def process_empty_name(data_info: dict, result_map):
    data_current_file = os.path.split(data_info['location']['location_path'])[1]
    if data_info['kind'] == 'ENUM_DECL' and 'members' in data_info and data_current_file in data_info['type']:
        for element in data_info['members']:
            result_map.setdefault(f'{data_current_file}-{element["name"]}', element)
    elif data_info['kind'] == 'ENUM_DECL' and 'members' in data_info and (data_current_file not in data_info['type']):
        result_map.setdefault(f'{data_current_file}-{data_info["type"]}', data_info)
    elif (data_info['kind'] == 'STRUCT_DECL' or data_info['kind'] == 'UNION_DECL') and \
            (data_current_file not in data_info['type']):
        result_map.setdefault(f'{data_current_file}-{data_info["type"]}', data_info)


def parse_file_result_by_child(result_map, root_node):
    children_list = root_node['children']
    for children in children_list:
        if children["name"] == '':
            process_empty_name(children, result_map)
            continue
        result_map.setdefault(f'{children["name"]}-{children["kind"]}', children)
    del root_node['children']
